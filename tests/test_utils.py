# test_link_manager.py
import csv

import openpyxl
import pytest

from django.core.paginator import Page

from shortener.models import Link
from shortener.utils import create_csv_file
from shortener.utils import create_xlsx_file
from shortener.utils import generate_short_hash
from shortener.utils import get_paginated_links
from shortener.utils import get_sorted_links
from shortener.utils import save_url_mapping


def test_generate_short_hash():
    # Проверяет генерацию хеша со стандартной длиной
    short_hash = generate_short_hash()
    assert len(short_hash) == 6
    # Проверяет генерацию хеша с кастомной длиной
    short_hash = generate_short_hash(length=10)
    assert len(short_hash) == 10


@pytest.mark.django_db
def test_save_url_mapping():
    url = 'https://www.google.com'
    short_hash = save_url_mapping(url)
    # Проверяет, является ли маппинг корректным
    link = Link.objects.get(short_hash=short_hash)
    assert link.original_url == url
    assert link.short_hash == short_hash


@pytest.mark.django_db
def test_get_sorted_links():
    # Проверяет соритровку по количеству переходов
    links = get_sorted_links()
    assert links == Link.objects.all().order_by('views_counter')


@pytest.mark.django_db
def test_get_paginated_links():
    links = Link.objects.all()
    page = get_paginated_links(links)
    # Проверяет, является ли возвращаемый объект Page
    assert isinstance(page, Page)
    # Проверяет пагинацию с кастомным размером страницы
    page = get_paginated_links(links, page_size=5)
    assert page.paginator.per_page == 5
    # Проверяет пагинацию с кастомным количеством страниц
    page = get_paginated_links(links, page_number=2)
    assert page.number == 2


@pytest.mark.django_db
def test_create_csv_file(links):
    create_csv_file(links)
    with open("statistics.csv") as file:
        reader = csv.reader(file)
        header = next(reader)
        assert header == [
            'Оригинальная ссылка',
            'Хеш ссылки',
            'Количество переходов',
        ]
        for i, row in enumerate(reader):
            assert row == [
                links[i].original_url,
                links[i].short_hash,
                links[i].views_counter,
            ]


@pytest.mark.django_db
def test_create_xlsx_file(data):
    create_xlsx_file(data)
    wb = openpyxl.load_workbook("statistics.xlsx")
    ws = wb.active
    header_row = [
        'id',
        'Оригинальная ссылка',
        'Хеш ссылки',
        'Количество переходов',
    ]
    for i, header_title in enumerate(header_row):
        assert ws.cell(row=1, column=i + 1).value == header_title
    for i, data_row in enumerate(data, 2):
        assert ws.cell(row=i, column=1).value == data_row.id
        assert ws.cell(row=i, column=2).value == data_row.original_url
        assert ws.cell(row=i, column=3).value == data_row.short_hash
        assert ws.cell(row=i, column=4).value == data_row.views_counter
