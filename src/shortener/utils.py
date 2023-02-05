import csv
import random
import string

from typing import List

import openpyxl

from django.core.paginator import Page
from django.core.paginator import Paginator

from .models import Link


def generate_short_hash(length: int = 6) -> str:
    """
    Генерирует случайный хеш для сокращенной ссылки

    :param length: Длина хеша
    :return: Строка размером length, которая может быть путем к сокращенной ссылке
    """
    return ''.join(
        random.choices(string.ascii_letters + string.digits, k=length)
    )


def save_url_mapping(url: str) -> str:
    """
    Сохраняет маппинг хеша к оригинальной ссылке

    :param url: Оригинальная ссылка.
    :return: Хеш для доступа к оригинальной ссылке.
    """
    while True:
        short_hash = generate_short_hash()
        if not Link.objects.filter(short_hash=short_hash).exists():
            break
    Link.objects.create(original_url=url, short_hash=short_hash)
    return short_hash


def get_sorted_links(sort_by: str = 'views_counter') -> List[Link]:
    """
    Получает все объекты типа "Link", отсортированные в заданной последовательности

    :param sort_by: Поле, задающее последовательность сортировки объектов
    :return: list объектов типа "Link", отсортированных в заданной последовательности
    """
    links = Link.objects.all().order_by(sort_by)
    return links


def get_paginated_links(
    links, page_size: int = 10, page_number: int = 1
) -> Page:
    """
    Получает все объекты типа "Link" после пагинации

    :param links: Список объектов типа "Link" подлежащих пагинации
    :param page_size: Количество объектов для показа на странице
    :param page_number: Количество страниц для показа на странице
    :return: Объект типа "Page", содержащий ссылки для показа на определенной странице
    """
    paginator = Paginator(links, page_size)
    links = paginator.get_page(page_number)
    return links


def create_csv_file(links: List[Link]) -> None:
    """
    Генерирует .csv файл со статистикой переходов по сокращенным сылкам

    :param links: Список ссылок для включения в .csv файл
    """
    with open('statistics.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(
            ['Оригинальная ссылка', 'Хеш ссылки', 'Количество переходов']
        )
        for link in links:
            writer.writerow(
                [link.original_url, link.short_hash, link.views_counter]
            )


def create_xlsx_file(data: List[Link]) -> None:
    """
    Генерирует .xlsx файл со статистикой переходов по сокращенным сылкам

    :param data: Список ссылок для включения в .xlsx файл
    """

    wb = openpyxl.Workbook()
    ws = wb.active

    header_row = [
        'id',
        'Оригинальная ссылка',
        'Хеш ссылки',
        'Количество переходов',
    ]
    for col_num, header_title in enumerate(header_row, 1):
        ws.cell(row=1, column=col_num, value=header_title)

    for row_num, data_row in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=data_row.id)
        ws.cell(row=row_num, column=2, value=data_row.original_url)
        ws.cell(row=row_num, column=3, value=data_row.short_hash)
        ws.cell(row=row_num, column=4, value=data_row.views_counter)

    wb.save("statistics.xlsx")
